"""Step 0 of the projection overhaul: reproduce the board-vs-consensus comparison.

Reads the FantasyPros draft sheet's position tabs (consensus stat line per
player, then a "high" and a "low" expert line), scores the AVG line in the
league's own settings, scales it to the board's games convention, joins to
the league's tiers csv by normalised name + position, and reports per
position: Spearman rank correlation, mean bias (board minus sheet), the ten
largest disagreements, and the deep-rank band table.

This is the acceptance test for item 1: the deep tail gap should close and
the QB rank correlation should rise.

Conventions (stated, not assumed):
  * The sheet's stat lines are full-season lines (17 games). The board's
    proj_pts are per-game model points x expected_games (16 by config). The
    sheet is scaled by --games/17 (default 16) so the two are on one basis.
  * Scoring comes from the league yaml's expected.scoring; stat keys the
    sheet has no column for (e.g. pass_td_40p) are ignored, as everywhere.
  * One workbook serves both leagues: the stat lines are format-free; only
    the sheet's own Aggregate/FLEX/RISK tabs are configured per league and
    those are not used here.

    venv\\Scripts\\python.exe scripts\\sheet_compare.py --league keefamania
    venv\\Scripts\\python.exe scripts\\sheet_compare.py --league omnibeta
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import polars as pl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from draftkit.config import Config  # noqa: E402
from draftkit.ids import normalize_name  # noqa: E402
from draftkit.seasondata import score_projection  # noqa: E402

SHEET = ROOT / "data" / "external" / "DraftSheets_2026_Keefamania_10tm_halfPPR_1flex.xlsx"
SHEET_GAMES = 17.0

# Column layout of each position tab (0-based, after Player, Team). The
# header names repeat ("YDS" twice), so the mapping is positional.
TAB_COLS = {
    "QB": ["pass_att", "pass_cmp", "pass_yd", "pass_td", "pass_int",
           "rush_att", "rush_yd", "rush_td", "fum_lost"],
    "RB": ["rush_att", "rush_yd", "rush_td", "rec", "rec_yd", "rec_td", "fum_lost"],
    "WR": ["rec", "rec_yd", "rec_td", "rush_att", "rush_yd", "rush_td", "fum_lost"],
    "TE": ["rec", "rec_yd", "rec_td", "fum_lost"],
}


def parse_tab(rows: list[tuple], pos: str) -> list[dict]:
    """rows: the tab's rows as tuples (header first). A player row carries the
    name; the next rows labelled 'high' / 'low' in the Team column are the
    expert extremes for that player. Returns [{name, team, avg, high, low}]
    where each of avg/high/low is a stat dict keyed like Sleeper's."""
    cols = TAB_COLS[pos]
    out: list[dict] = []
    cur: dict | None = None
    for r in rows[1:]:
        name, team = (r[0] if len(r) > 0 else None), (r[1] if len(r) > 1 else None)
        stats = {k: r[2 + i] for i, k in enumerate(cols)
                 if len(r) > 2 + i and isinstance(r[2 + i], (int, float))}
        if isinstance(name, str) and name.strip() and name.strip() != "\xa0":
            cur = {"name": name.strip(), "team": team, "avg": stats, "high": None, "low": None}
            out.append(cur)
        elif cur is not None and isinstance(team, str) and team.strip().lower() in ("high", "low"):
            cur[team.strip().lower()] = stats
    return out


def read_sheet(path: Path) -> dict[str, list[dict]]:
    import openpyxl  # local import: only this script needs it
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for pos in TAB_COLS:
        ws = wb[pos]
        out[pos] = parse_tab(list(ws.iter_rows(values_only=True)), pos)
    return out


def scoring_from_league(cfg: Config) -> dict[str, float]:
    block = cfg.get("scoring") or (cfg.get("expected") or {}).get("scoring") or {}
    if not block:
        raise SystemExit("league yaml carries no scoring block")
    return {k: float(v) for k, v in block.items()}


def spearman(a: list[float], b: list[float]) -> float:
    """Rank correlation with average ranks for ties."""
    def ranks(x):
        order = sorted(range(len(x)), key=lambda i: x[i])
        r = [0.0] * len(x)
        i = 0
        while i < len(order):
            j = i
            while j + 1 < len(order) and x[order[j + 1]] == x[order[i]]:
                j += 1
            avg = (i + j) / 2 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r
    n = len(a)
    if n < 3:
        return float("nan")
    ra, rb = ranks(a), ranks(b)
    ma, mb = sum(ra) / n, sum(rb) / n
    cov = sum((x - ma) * (y - mb) for x, y in zip(ra, rb))
    va = sum((x - ma) ** 2 for x in ra) ** 0.5
    vb = sum((y - mb) ** 2 for y in rb) ** 0.5
    return cov / (va * vb) if va and vb else float("nan")


def compare(sheet: dict[str, list[dict]], board: pl.DataFrame, scoring: dict,
            games: float, column: str = "proj_pts") -> dict:
    """Join and measure. Returns {pos: {...}} plus 'unmatched'. `column` is
    the board column under test (proj_pts, or a parallel source such as
    proj_consensus_pts); rows where it is null are left out of the join."""
    scale = games / SHEET_GAMES
    result: dict = {}
    for pos, players in sheet.items():
        sub = (board.filter((pl.col("pos") == pos) & pl.col(column).is_not_null())
               .sort(column, descending=True))
        bmap = {}
        for i, (name, pts) in enumerate(sub.select(["player", column]).iter_rows(), 1):
            bmap.setdefault(normalize_name(name), (i, float(pts), name))
        rows = []
        unmatched = []
        ranked = sorted(players, key=lambda p: -score_projection(p["avg"], scoring))
        for srank, p in enumerate(ranked, 1):
            key = normalize_name(p["name"])
            s_pts = score_projection(p["avg"], scoring) * scale
            hi = score_projection(p["high"], scoring) * scale if p["high"] else None
            lo = score_projection(p["low"], scoring) * scale if p["low"] else None
            if key not in bmap:
                unmatched.append(p["name"])
                continue
            brank, bpts, bname = bmap[key]
            rows.append({"name": bname, "sheet_rank": srank, "sheet_pts": s_pts,
                         "board_rank": brank, "board_pts": bpts,
                         "sheet_high": hi, "sheet_low": lo,
                         "pts_diff": bpts - s_pts, "rank_diff": brank - srank})
        top = [r for r in rows if r["sheet_rank"] <= 36]
        bands = []
        for lo_, hi_ in ((1, 12), (13, 24), (25, 36), (37, 48), (49, 60), (61, 80)):
            band = [r for r in rows if lo_ <= r["sheet_rank"] <= hi_]
            if band:
                bands.append({"band": f"{lo_}-{hi_}", "n": len(band),
                              "sheet": sum(r["sheet_pts"] for r in band) / len(band),
                              "board": sum(r["board_pts"] for r in band) / len(band),
                              "diff": sum(r["pts_diff"] for r in band) / len(band)})
        result[pos] = {
            "matched": len(rows), "unmatched": unmatched,
            "spearman_all": spearman([r["sheet_rank"] for r in rows], [r["board_rank"] for r in rows]),
            "spearman_top36": spearman([r["sheet_rank"] for r in top], [r["board_rank"] for r in top]),
            "bias_all": (sum(r["pts_diff"] for r in rows) / len(rows)) if rows else float("nan"),
            "bias_top36": (sum(r["pts_diff"] for r in top) / len(top)) if top else float("nan"),
            "top_rank_disagreements": sorted(top, key=lambda r: -abs(r["rank_diff"]))[:10],
            "top_pts_disagreements": sorted(rows, key=lambda r: -abs(r["pts_diff"]))[:10],
            "bands": bands,
            "rows": rows,
        }
    return result


def render(league: str, games: float, res: dict, scoring: dict | None = None,
           column: str = "proj_pts") -> str:
    sc = ", ".join(f"{k} {v:g}" for k, v in (scoring or {}).items())
    L = [f"# Board vs FantasyPros consensus — {league}",
         "",
         "## Scoring basis (read this before comparing with any other sheet-vs-board number)",
         "",
         f"- Sheet side: the position tabs' **AVG stat lines** (raw consensus lines, NOT the "
         f"Aggregate tab, which already carries the sheet's missed-games adjustment), scored "
         f"with the league yaml's scoring: {sc or 'n/a'}. Stat keys the sheet has no column "
         f"for (e.g. pass_td_40p) are ignored.",
         f"- Games: sheet lines are {SHEET_GAMES:g}-game season totals, scaled by "
         f"{games:g}/{SHEET_GAMES:g} to the board's `expected_games` basis. No injury or "
         f"missed-games adjustment is applied on either side.",
         f"- Board side: `{column}` from the league's tiers csv. Bias = board minus sheet.",
         "- Ranks: within position, by the respective points. Spearman over matched players; "
         "'top 36' restricts to the sheet's top 36 at the position.",
         "",
         "A uniform negative bias across a position's starters cancels in VORP and is not a "
         "defect. Differences BETWEEN positions' biases do not cancel and are worth noting.",
         ""]
    L += ["| pos | matched | Spearman (all) | Spearman (sheet top 36) | bias all | bias top 36 | unmatched |",
          "|---|---|---|---|---|---|---|"]
    for pos, r in res.items():
        L.append(f"| {pos} | {r['matched']} | {r['spearman_all']:.2f} | {r['spearman_top36']:.2f} | "
                 f"{r['bias_all']:+.1f} | {r['bias_top36']:+.1f} | {len(r['unmatched'])} |")
    for pos, r in res.items():
        L += ["", f"## {pos}", "", "Deep-rank bands (by sheet rank):", "",
              "| band | n | sheet | board | diff |", "|---|---|---|---|---|"]
        for b in r["bands"]:
            L.append(f"| {b['band']} | {b['n']} | {b['sheet']:.0f} | {b['board']:.0f} | {b['diff']:+.0f} |")
        L += ["", "Largest rank disagreements (sheet top 36):", "",
              "| player | sheet rk | board rk | sheet pts | board pts |", "|---|---|---|---|---|"]
        for x in r["top_rank_disagreements"]:
            L.append(f"| {x['name']} | {x['sheet_rank']} | {x['board_rank']} | {x['sheet_pts']:.0f} | {x['board_pts']:.0f} |")
        L += ["", "Largest point disagreements (all matched):", "",
              "| player | sheet rk | board rk | sheet pts | board pts | diff |", "|---|---|---|---|---|---|"]
        for x in r["top_pts_disagreements"]:
            L.append(f"| {x['name']} | {x['sheet_rank']} | {x['board_rank']} | {x['sheet_pts']:.0f} | "
                     f"{x['board_pts']:.0f} | {x['pts_diff']:+.0f} |")
        if r["unmatched"]:
            L += ["", f"Sheet players not on the board ({len(r['unmatched'])}): "
                  + ", ".join(r["unmatched"][:25]) + (" …" if len(r["unmatched"]) > 25 else "")]
    return "\n".join(L) + "\n"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--league", required=True)
    ap.add_argument("--sheet", default=str(SHEET))
    ap.add_argument("--games", type=float, default=None,
                    help="games basis for the sheet (default: config projections.expected_games)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--column", default="proj_pts",
                    help="board column under test (default proj_pts; e.g. proj_consensus_pts)")
    ap.add_argument("--board", default=None, help="csv to grade instead of the league tiers csv")
    a = ap.parse_args()

    cfg = Config.load(league=a.league)
    games = a.games if a.games is not None else float((cfg.get("projections") or {}).get("expected_games", 16.0))
    board_path = Path(a.board) if a.board else cfg.scoped(cfg.root / "tiers.csv")
    board = pl.read_csv(board_path, infer_schema_length=2000)
    if a.column not in board.columns:
        raise SystemExit(f"{board_path.name} has no column {a.column}")
    sheet = read_sheet(Path(a.sheet))
    scoring = scoring_from_league(cfg)
    res = compare(sheet, board, scoring, games, column=a.column)
    md = render(a.league, games, res, scoring, column=a.column)
    suffix = "" if a.column == "proj_pts" else f".{a.column}"
    out = Path(a.out) if a.out else ROOT / "reports" / f"sheet_compare.{a.league}{suffix}.md"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(md, encoding="utf-8")
    for pos, r in res.items():
        print(f"{pos}: matched {r['matched']:3}  spearman all {r['spearman_all']:.2f}  top36 {r['spearman_top36']:.2f}"
              f"  bias all {r['bias_all']:+.1f}  top36 {r['bias_top36']:+.1f}  unmatched {len(r['unmatched'])}")
        print("   bands: " + "  ".join(f"{b['band']}:{b['diff']:+.0f}" for b in r["bands"]))
    print(f"-> {out}")


if __name__ == "__main__":
    main()
