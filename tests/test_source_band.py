"""A source's OWN stated range, kept distinct from disagreement BETWEEN sources.

The FantasyPros sheet publishes a high and a low line either side of each
player's base line, and its own Aggregate tab averages the three. The loader
skipped the extremes, so `proj_sd` was zero for every sheet player and the
late-round dispersion objective could never fire on a single-source board.

Those two quantities are not the same thing:

  proj_sd    two forecasters disagreeing. Zero at one source is an ABSENCE of
             evidence, not a narrow forecast, which is why the n >= 2 guard
             exists and stays.
  proj_band  one forecaster stating his own uncertainty. Real evidence, and
             available at n = 1.

They get separate columns and are never summed.
"""

from __future__ import annotations

import polars as pl
import pytest

from draftkit import external as X
from draftkit.tracker import Tracker

SCORING = {"rush_yd": 0.1, "rush_td": 6.0, "rec": 0.5, "rec_yd": 0.1,
           "rec_td": 6.0, "fum_lost": -2.0, "rush_att": 0.0}

HDR = ("Player", "Team", "ATT", "YDS", "TDS", "REC", "YDS", "TDS", "FL")


def _rb_rows():
    """The sheet's real shape: a named base row, then unnamed high and low."""
    return [
        HDR,
        ("\xa0", None, None, None, None, None, None, None, None),      # spacer
        ("Base Guy", "DET", 200.0, 1000.0, 10.0, 50.0, 400.0, 2.0, 1.0),
        (None, "high", 220.0, 1200.0, 13.0, 60.0, 500.0, 3.0, 1.0),
        (None, "low", 180.0, 800.0, 7.0, 40.0, 300.0, 1.0, 1.0),
        ("No Range Guy", "ATL", 100.0, 500.0, 4.0, 20.0, 150.0, 1.0, 0.0),
    ]


# ------------------------------------------------------------------- parsing

def test_the_high_and_low_rows_attach_to_the_player_above_them():
    out = X.parse_sheet_tab(_rb_rows(), "RB")
    assert [p["name"] for p in out] == ["Base Guy", "No Range Guy"]
    assert out[0]["line_hi"]["rush_yd"] == 1200.0
    assert out[0]["line_lo"]["rush_yd"] == 800.0
    assert out[1]["line_hi"] is None and out[1]["line_lo"] is None


def test_the_spacer_row_is_still_not_a_player():
    """A non-breaking space in the name column, and the mojibake it reads back
    as, are the sheet's blank separators."""
    for blank in ("\xa0", "Â\xa0", "   "):
        rows = [HDR, (blank, None, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0)]
        assert X.parse_sheet_tab(rows, "RB") == []


def test_an_orphan_high_row_is_dropped_not_attached_to_nothing():
    rows = [HDR, (None, "high", 220.0, 1200.0, 13.0, 60.0, 500.0, 3.0, 1.0)]
    assert X.parse_sheet_tab(rows, "RB") == []


# ------------------------------------------------------------------ the band

class _Index:
    def match(self, name, pos, team):
        return {"Base Guy": "1", "No Range Guy": "2"}.get(name)


def _sheet_frame(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for pos in ("QB", "RB", "WR", "TE"):
        ws = wb.create_sheet(pos)
        for r in (_rb_rows() if pos == "RB" else [HDR]):
            ws.append(list(r))
    p = tmp_path / "sheet.xlsx"
    wb.save(p)
    return X.from_sheet(p, SCORING, _Index(), as_of="2026-09-01")[0]


def test_the_band_is_the_sd_of_low_base_and_high(tmp_path):
    f = _sheet_frame(tmp_path)
    row = {r["sleeper_id"]: r for r in f.iter_rows(named=True)}["1"]
    # the expectation, computed the way a reader would: score all three lines
    trio = [X.score_projection(ln, SCORING) for ln in (
        {"rush_yd": 800.0, "rush_td": 7.0, "rec": 40.0, "rec_yd": 300.0, "rec_td": 1.0, "fum_lost": 1.0, "rush_att": 180.0},
        {"rush_yd": 1000.0, "rush_td": 10.0, "rec": 50.0, "rec_yd": 400.0, "rec_td": 2.0, "fum_lost": 1.0, "rush_att": 200.0},
        {"rush_yd": 1200.0, "rush_td": 13.0, "rec": 60.0, "rec_yd": 500.0, "rec_td": 3.0, "fum_lost": 1.0, "rush_att": 220.0})]
    mu = sum(trio) / 3
    want = (sum((x - mu) ** 2 for x in trio) / 3) ** 0.5
    assert row["pts17_band"] == pytest.approx(want)
    assert row["pts17_band"] > 0


def test_the_point_estimate_is_still_the_BASE_line_not_the_mean(tmp_path):
    """Averaging the trio into pts17 is a SEPARATE change. Loading the band
    must not move a single projection, which is what makes the two
    independently gradeable."""
    f = _sheet_frame(tmp_path)
    row = {r["sleeper_id"]: r for r in f.iter_rows(named=True)}["1"]
    base = X.score_projection(
        {"rush_att": 200.0, "rush_yd": 1000.0, "rush_td": 10.0, "rec": 50.0,
         "rec_yd": 400.0, "rec_td": 2.0, "fum_lost": 1.0}, SCORING)
    assert row["pts17"] == pytest.approx(base)


def test_a_player_without_a_range_gets_a_null_band_not_a_zero(tmp_path):
    """Zero would read as 'this forecast is certain'. It means 'not stated'."""
    f = _sheet_frame(tmp_path)
    row = {r["sleeper_id"]: r for r in f.iter_rows(named=True)}["2"]
    assert row["pts17_band"] is None


# ------------------------------------------------------------------- schema

def test_combine_keeps_the_canonical_column_order_with_the_band(tmp_path):
    f = _sheet_frame(tmp_path)
    out = X.combine([f])
    assert list(out.columns) == list(X.SCHEMA_COMBINED)
    assert "pts17_band" in X.DISPERSION


def test_one_source_still_reports_no_cross_source_disagreement(tmp_path):
    """The band must not leak into proj_sd. n_sources stays 1 and the sd
    stays 0, because one source has not disagreed with anybody."""
    out = X.combine([_sheet_frame(tmp_path)])
    row = {r["sleeper_id"]: r for r in out.iter_rows(named=True)}["1"]
    assert row["n_sources"] == 1 and row["pts17_sd"] == 0.0
    assert row["pts17_band"] > 0


# -------------------------------------------------------- the engine's knob

def _t(**kw):
    t = object.__new__(Tracker)
    t.late_round_dispersion = kw.pop("on", True)
    for k, v in kw.items():
        setattr(t, k, v)
    return t


def test_the_knob_off_ignores_both_quantities():
    t = _t(on=False)
    assert t._dispersion_for({"proj_sd": 9.0, "n_sources": 4, "proj_band": 7.0}) is None


def test_cross_source_disagreement_wins_when_it_exists():
    """Preferring proj_sd keeps the arm identical to the one plan A3 was
    written and pre-registered for."""
    t = _t()
    assert t._dispersion_for({"proj_sd": 9.0, "n_sources": 4, "proj_band": 7.0}) == 9.0


def test_a_single_source_falls_back_to_its_own_band():
    t = _t()
    assert t._dispersion_for({"proj_sd": 0.0, "n_sources": 1, "proj_band": 7.0}) == 7.0


def test_a_single_source_with_no_band_still_gets_nothing():
    """This is the pre-change behaviour and it must survive: an sd of zero at
    one source is an absence of evidence, never a narrow forecast."""
    t = _t()
    assert t._dispersion_for({"proj_sd": 0.0, "n_sources": 1, "proj_band": None}) is None
    assert t._dispersion_for({"proj_sd": 0.0, "n_sources": 1, "proj_band": 0.0}) is None
    assert t._dispersion_for({"proj_sd": None, "n_sources": 0}) is None


def test_the_two_quantities_are_never_summed():
    """Combining them in quadrature is a modelling claim nobody has measured."""
    t = _t()
    got = t._dispersion_for({"proj_sd": 3.0, "n_sources": 2, "proj_band": 4.0})
    assert got == 3.0 and got != 5.0 and got != 7.0


def test_every_board_loader_carries_the_same_engine_fields():
    """The band has to survive tiers.csv -> the engine, by BOTH routes.

    The offline replays go through engine_parity.load_board and the live Yahoo
    rig through yahoo_bridge.load_players. They were two hand-written copies of
    one list and had already drifted: the live rig never carried proj_band, so
    an offline replay would have shown the analyst spread working while a real
    room silently ignored it. An offline test alone cannot catch that, which is
    why this asserts the shared definition rather than either loader's source.
    """
    from draftkit.boardrow import ENGINE_FIELDS, engine_fields

    for required in ("proj_band", "proj_sd", "proj_market_pts", "n_sources", "adp"):
        assert required in ENGINE_FIELDS

    row = {"player": "X", "pos": "RB", "vorp": "1.0", "proj_pts": "100.0",
           "proj_band": "12.5", "proj_sd": "", "n_sources": "1", "adp": "45"}
    got = engine_fields(row)
    assert got["proj_band"] == 12.5
    assert got["proj_sd"] is None, "an empty cell is absence, not zero"
    assert got["adp"] == 45.0 and got["vorp_flex"] == 1.0

    # and neither loader may reintroduce a private copy of the list
    import inspect
    import importlib.util
    from pathlib import Path
    root = Path(__file__).resolve().parents[1]
    for name, fn in (("engine_parity", "load_board"), ("yahoo_bridge", "load_players")):
        spec = importlib.util.spec_from_file_location(name, root / "scripts" / f"{name}.py")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        src = inspect.getsource(getattr(mod, fn))
        assert "engine_fields(r)" in src, f"{name}.{fn} stopped using the shared fields"
        assert '"proj_sd"' not in src, f"{name}.{fn} grew its own copy again"


def test_external_projection_selects_every_dispersion_column():
    """The regression that hid the band: external_projection kept its own
    hardcoded copy of the dispersion column list."""
    import inspect

    from draftkit import projections
    src = inspect.getsource(projections.external_projection)
    assert "X.DISPERSION" in src
    assert '("n_sources", "pts17_sd", "pts17_hi", "pts17_lo")' not in src
