"""draftkit must reproduce the spreadsheet's own numbers, given league scoring.

The workbook caches Excel's computed result for every cell, so this is a diff
against ground truth rather than against anyone's reading of the formulas. It
is the check that makes "the projection is an INPUT" (DECISIONS #21) an
auditable claim instead of a hope.

It caught two real gaps:

  * the loader read the base stat line only, while the sheet averages the
    analyst panel's low, base and high (now `pts17_band`);
  * the sheet adds a per-position ROOKIE BUMP that the loader dropped, worth
    up to 65 season points and applied to 42 players.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from draftkit import external as X
from draftkit.config import Config

ROOT = Path(__file__).resolve().parents[1]
SHEET = ROOT / "data/external/DraftSheets_2026_Keefamania_10tm_halfPPR_1flex.xlsx"

# the sheet's own AVG cell per tab (1-based column): the base line scored with
# the Scoring tab, plus the tab's rookie bump
AVG_COL = {"QB": 16, "RB": 14, "WR": 14, "TE": 10}

pytestmark = pytest.mark.skipif(not SHEET.exists(), reason="draft sheet not present")


class _PassThrough:
    """Every name resolves to itself, so the diff covers the whole sheet and
    not only the players Sleeper happens to know."""

    def match(self, name, pos, team):
        return f"{pos}:{name}"


@pytest.fixture(scope="module")
def loaded():
    cfg = Config.load(league="keefamania")
    scoring = cfg.get("scoring") or (cfg.get("expected") or {}).get("scoring") or {}
    frame, unmatched = X.from_sheet(SHEET, scoring, _PassThrough(), as_of="test")
    return frame, unmatched, scoring


@pytest.fixture(scope="module")
def cached():
    import openpyxl
    wb = openpyxl.load_workbook(SHEET, data_only=True)
    out = {}
    for pos, col in AVG_COL.items():
        for r in list(wb[pos].iter_rows(values_only=True))[1:]:
            n = r[0]
            if not (isinstance(n, str) and n.replace("\xa0", "").replace("Â", "").strip()):
                continue
            v = r[col - 1] if len(r) >= col else None
            if isinstance(v, (int, float)):
                out[f"{pos}:{n.strip()}"] = float(v)
    return out


def test_every_projection_matches_the_spreadsheets_own_value(loaded, cached):
    """The headline. A mismatch means draftkit and the spreadsheet disagree
    about what the same stat line is worth in this league."""
    frame, _unmatched, _ = loaded
    mine = {r["sleeper_id"]: r["pts17"] for r in frame.iter_rows(named=True)}
    compared, bad = 0, []
    for key, theirs in cached.items():
        if key not in mine:
            continue
        compared += 1
        if abs(mine[key] - theirs) >= 0.01:
            bad.append((abs(mine[key] - theirs), key, mine[key], theirs))
    assert compared >= 480, f"only {compared} players compared; the tabs moved"
    bad.sort(reverse=True)
    assert not bad, (
        f"{len(bad)} of {compared} disagree with the spreadsheet. Worst:\n"
        + "\n".join(f"  {k}: draftkit {a:.2f} vs sheet {b:.2f}" for _d, k, a, b in bad[:10]))


def test_the_league_scoring_matches_the_sheets_own_inputs():
    """A silent scoring drift would make every projection wrong together, so
    the diff above would still pass while both were wrong."""
    import openpyxl
    cfg = Config.load(league="keefamania")
    scoring = cfg.get("scoring") or (cfg.get("expected") or {}).get("scoring") or {}
    ws = openpyxl.load_workbook(SHEET, data_only=True)["Scoring"]
    sheet_in = {}
    for r in ws.iter_rows(min_row=1, max_row=26, values_only=True):
        if isinstance(r[0], str) and isinstance(r[1], (int, float)):
            sheet_in[r[0].strip()] = float(r[1])

    # the sheet states yards PER POINT; the league yaml states points per yard
    assert scoring["pass_yd"] == pytest.approx(1.0 / sheet_in["PassYDS"])
    assert scoring["rush_yd"] == pytest.approx(1.0 / sheet_in["RushYDS"])
    assert scoring["rec_yd"] == pytest.approx(1.0 / sheet_in["RecYDS"])
    assert scoring["pass_td"] == pytest.approx(sheet_in["PassTDs"])
    assert scoring["rush_td"] == pytest.approx(sheet_in["RushTDS"])
    assert scoring["rec_td"] == pytest.approx(sheet_in["RecTDS"])
    assert scoring["pass_int"] == pytest.approx(sheet_in["INTS"])
    assert scoring["fum_lost"] == pytest.approx(sheet_in["FL"])
    assert scoring["rec"] == pytest.approx(sheet_in["RB PPR"]) == pytest.approx(sheet_in["WR PPR"])

    # terms the sheet scores and draftkit does not: all zero in this league,
    # which is WHY the diff passes. A league that pays for them would need
    # first downs and completions carried through the stat line.
    for k in ("PassCMP", "PassINCMP", "SACKS", "RushATT", "Pass1D", "Rush1D", "Rec1D"):
        assert sheet_in[k] == 0.0, (
            f"the sheet scores {k} at {sheet_in[k]} and draftkit's stat line does "
            "not carry it; parity above is accidental until it does")


def test_the_rookie_bump_is_applied_and_is_large(loaded, cached):
    """Not applying it under-projects rookies by up to 65 points against the
    sheet. This asserts the size so a silent regression to 'no bump' fails
    loudly rather than shifting the deep board by tens of points."""
    import openpyxl
    wb_v = openpyxl.load_workbook(SHEET, data_only=True)
    wb_f = openpyxl.load_workbook(SHEET, data_only=False)
    total, n = 0.0, 0
    for pos in ("RB", "WR"):
        col = X.sheet_bump_column(wb_f[pos])
        assert col is not None, f"{pos}: the bump column moved"
        for r in list(wb_v[pos].iter_rows(values_only=True))[1:]:
            nm = r[0]
            if not (isinstance(nm, str) and nm.replace("\xa0", "").replace("Â", "").strip()):
                continue
            b = r[col] if len(r) > col else None
            if isinstance(b, (int, float)) and b > 0:
                total += float(b)
                n += 1
    assert n >= 40, f"only {n} bumped players found; the sheet's rookie set changed"
    assert total / n > 30.0, "the mean bump collapsed; check the coefficients"


def test_qb_and_te_have_no_bump_column():
    """The bump is RB and WR only. Finding one at QB or TE would mean the
    formula-shape locator is matching something else."""
    import openpyxl
    wb_f = openpyxl.load_workbook(SHEET, data_only=False)
    for pos in ("QB", "TE"):
        assert X.sheet_bump_column(wb_f[pos]) is None


def test_the_band_is_carried_and_zero_means_the_panel_agreed(loaded):
    """Deep players get high and low rows that just repeat the base line, so a
    band of exactly zero is the sheet saying "no spread published for him",
    not a missing column. Both read as no evidence downstream."""
    frame, _u, _s = loaded
    rows = list(frame.iter_rows(named=True))
    banded = [r for r in rows if r["pts17_band"] is not None]
    assert len(banded) == len(rows), "the high/low rows stopped being attached"
    assert all(r["pts17_band"] >= 0 for r in banded)
    spread = [r for r in banded if r["pts17_band"] > 0]
    assert len(spread) >= 300, (
        f"only {len(spread)} players carry a real spread; the panel's high/low "
        "rows are probably no longer being read")


def test_nothing_is_dropped_when_every_name_resolves(loaded):
    _f, unmatched, _s = loaded
    assert unmatched == []
