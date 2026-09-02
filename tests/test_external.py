"""Projections as an input (DECISIONS 2026-09-02 #21).

Two sources into one schema, first source wins per player, one games
scaling at the end, and the one tail rule: non-starters project zero.
"""

from __future__ import annotations

import json

import polars as pl

from draftkit import external as X

HALF = {"rec": 0.5, "pass_yd": 0.04, "pass_td": 4.0, "pass_int": -1.0,
        "rush_yd": 0.1, "rec_yd": 0.1, "rush_td": 6.0, "rec_td": 6.0, "fum_lost": -2.0}


class FakeIndex:
    def __init__(self, table):
        self.table = table

    def match(self, name, pos, team):
        return self.table.get((name, pos))


def test_sheet_tab_parses_consensus_rows_and_skips_expert_extremes():
    rows = [
        ("Player", "Team", "ATT", "YDS", "TDS", "REC", "YDS", "TDS", "FL", "FPTS"),
        ("\xa0", None, None, None, None, None, None, None, None, None),
        ("Jahmyr Gibbs", "DET", 275.2, 1383.7, 13.8, 71.3, 581.1, 4.1, 1.1, 337.4),
        (None, "high", 283.5, 1422, 15, 74, 625, 5, 1, 367.1),
        (None, "low", 260, 1300, 12, 65, 520, 3, 2, 300.0),
        ("Bijan Robinson", "ATL", 300, 1400, 12, 60, 500, 3, 1, 320.0),
    ]
    out = X.parse_sheet_tab(rows, "RB")
    assert [p["name"] for p in out] == ["Jahmyr Gibbs", "Bijan Robinson"]
    assert out[0]["line"]["rec_yd"] == 581.1 and out[0]["line"]["fum_lost"] == 1.1


def test_from_sheet_scores_in_league_settings_and_reports_unmatched(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for pos, hdr in (("QB", ["Player", "Team"] + ["x"] * 9), ("RB", ["Player", "Team"] + ["x"] * 7),
                     ("WR", ["Player", "Team"] + ["x"] * 7), ("TE", ["Player", "Team"] + ["x"] * 4)):
        ws = wb.create_sheet(pos)
        ws.append(hdr)
    wb["RB"].append(["Jahmyr Gibbs", "DET", 275.2, 1383.7, 13.8, 71.3, 581.1, 4.1, 1.1])
    wb["RB"].append([None, "high", 283.5, 1422, 15, 74, 625, 5, 1])
    wb["RB"].append(["Nobody Here", "FA", 10, 50, 0, 0, 0, 0, 0])
    p = tmp_path / "sheet.xlsx"
    wb.save(p)
    df, unmatched = X.from_sheet(p, HALF, FakeIndex({("Jahmyr Gibbs", "RB"): "4866"}), as_of="2026-09-01")
    assert unmatched == ["Nobody Here (RB)"]
    r = df.row(0, named=True)
    assert r["sleeper_id"] == "4866" and r["source"] == "fantasypros_sheet" and r["as_of"] == "2026-09-01"
    assert abs(r["pts17"] - 337.33) < 0.05        # the sheet's own half-PPR FPTS for that line
    assert json.loads(r["line"])["rush_yd"] == 1383.7


def test_from_sleeper_uses_the_common_schema_and_skips_placeholders(tmp_path):
    rows = [{"player_id": "4984", "team": "BUF", "updated_at": 1788249037361,
             "player": {"position": "QB", "first_name": "Josh", "last_name": "Allen"},
             "stats": {"gp": 18.0, "pass_yd": 3650.0, "pass_td": 28.0, "adp_half_ppr": 20.7}},
            {"player_id": "9", "team": None, "player": {"position": "QB"}, "stats": {"gp": 18.0, "adp_half_ppr": 600.0}}]

    def getter(url):
        return rows if url.endswith("=QB") else []

    df = X.from_sleeper(2026, HALF, tmp_path, getter=getter, ttl=3600)
    assert df.height == 1
    r = df.row(0, named=True)
    assert r["name"] == "Josh Allen" and r["source"] == "sleeper_rotowire" and r["as_of"] == "2026-09-01"
    assert abs(r["pts17"] - (3650 * 0.04 + 28 * 4)) < 1e-9
    assert "gp" not in json.loads(r["line"]) and "adp_half_ppr" not in json.loads(r["line"])


def test_combine_lets_the_first_source_win_and_later_ones_fill_gaps():
    a = pl.DataFrame({"sleeper_id": ["1", "2"], "name": ["A", "B"], "pos": ["RB", "RB"], "team": ["X", "X"],
                      "pts17": [200.0, 150.0], "source": ["sheet", "sheet"], "as_of": ["d", "d"], "line": ["{}", "{}"]},
                     schema=X.SCHEMA)
    b = pl.DataFrame({"sleeper_id": ["2", "3"], "name": ["B", "C"], "pos": ["RB", "RB"], "team": ["X", "X"],
                      "pts17": [999.0, 100.0], "source": ["sleeper", "sleeper"], "as_of": ["e", "e"], "line": ["{}", "{}"]},
                     schema=X.SCHEMA)
    out = X.combine([a, b])
    got = {r["sleeper_id"]: r for r in out.iter_rows(named=True)}
    assert got["2"]["pts17"] == 150.0 and got["2"]["source"] == "sheet"
    assert got["3"]["source"] == "sleeper" and out.height == 3
    assert X.combine([X.empty(), b]).height == 2


def test_non_starters_go_to_zero_only_when_depth_chart_and_market_agree():
    teams = 10
    df = pl.DataFrame({
        "sleeper_id": ["qb1", "qb2", "glitch", "nomkt", "wr_deep", "rb3", "hback"],
        "name": ["Starter", "Backup", "Glitch", "NoMarket", "DeepWR", "RB3", "HBack"],
        "pos": ["QB", "QB", "QB", "QB", "WR", "RB", "TE"],
        "team": ["A", "A", "B", "C", "D", "E", "F"],
        "ecr": [3.0, 40.0, 5.0, None, 90.0, None, None],   # rb3: depth 3 and unranked
        "adp": [None, None, None, None, None, None, None],
        "proj_pts": [300.0, 200.0, 280.0, 190.0, 120.0, 90.0, 60.0],
    }, schema_overrides={"adp": pl.Float64, "ecr": pl.Float64})
    filler = pl.DataFrame({"sleeper_id": [f"f{i}" for i in range(12)], "name": [f"F{i}" for i in range(12)],
                           "pos": ["QB"] * 12, "team": ["Z"] * 12, "ecr": [10.0 + i for i in range(12)],
                           "adp": [None] * 12, "proj_pts": [250.0] * 12}, schema_overrides={"adp": pl.Float64})
    df = pl.concat([df, filler], how="diagonal_relaxed")
    depth = pl.DataFrame({"sleeper_id": ["qb1", "qb2", "glitch", "nomkt", "wr_deep", "rb3", "hback"],
                          "depth_order": [1, 2, 2, 2, 9, 3, 6],
                          "depth_pos": ["QB", "QB", "QB", "QB", "RWR", "RB", "RB"]})
    out = X.zero_non_starters(df, depth, teams)
    g = {r["sleeper_id"]: r for r in out.iter_rows(named=True)}
    assert g["qb1"]["proj_pts"] == 300.0 and g["qb1"]["non_starter"] is False
    assert g["qb2"]["proj_pts"] == 0.0 and g["qb2"]["contingent_of"] == "Starter"
    assert g["glitch"]["proj_pts"] == 280.0, "market ranks him QB3: one source is not enough"
    assert g["nomkt"]["proj_pts"] == 0.0, "no ECR/ADP at all is the market's strongest 'backup'"
    assert g["wr_deep"]["proj_pts"] == 120.0, "WR chart is per slot; never zeroed on it"
    assert g["rb3"]["proj_pts"] == 0.0 and g["rb3"]["contingent_of"] is None, "no order-1 RB known on team E"
    assert g["hback"]["proj_pts"] == 60.0, "a TE filed under the RB chart is left alone"
    assert "_mkt_rank" not in out.columns
