"""Projections as an INPUT: external stat lines, scored in league settings.

DECISIONS 2026-09-02 #21 (simplification): the engine's edge is roster-aware
timing -- slots still needed, value against the slot a player fills, value
against what is freely available later, timing against the other rosters.
None of that needs our projections to beat consensus, and the FantasyPros
comparison showed they do not. So the projection layer stops being a model
and becomes an input. Everything downstream is unchanged.

Two sources, one schema:

    sleeper_id · name · pos · team · pts17 · source · as_of · line

  * pts17 is the stat line scored with the league yaml's scoring, as a
    17-game season total -- the convention every source is put on before the
    engine's own `projections.games` (16) scaling is applied ONCE, at the
    end, for everyone. Sleeper's `gp` (18) is a week count and is ignored.
  * `line` is the raw stat dict (Sleeper-style keys) as JSON, kept so the
    number can be audited back to its inputs.

Sources:
  sheet   the FantasyPros draft sheet's position tabs (data/external/
          DraftSheets_2026_*.xlsx, committed read-only). The consensus AVG
          line per player; the high/low expert lines are ignored here. Names
          are resolved to Sleeper ids through the same matcher the market
          table uses.
  sleeper https://api.sleeper.app/projections/nfl/<season>?season_type=
          regular&position[]=POS -- Rotowire's lines, refreshed regularly.
          The source for every draft after this one.

Precedence is the config's order (2026: sheet, then Sleeper for players the
sheet lacks). Every row says which source it came from and when.

Non-starters go to zero (the one tail rule): a player the depth chart lists
behind his position's starters, whom the market also does not rank as a
starter (or does not rank at all), projects 0 whatever his per-game talent.
`contingent_of` names the starter ahead of him so the informational handcuff
column still works. See DECISIONS #21 for the WR caveat.
"""

from __future__ import annotations

import datetime as dt
import json
import logging
import re
from pathlib import Path

import polars as pl

from .consensus import POSITIONS as SLEEPER_POSITIONS
from .consensus import ConsensusUnavailable, fetch_position
from .role import GATED, STARTERS, depth_orders
from .seasondata import score_projection

log = logging.getLogger("draftkit")

LINE_GAMES = 17.0
# Each source's own games convention, MEASURED 2026-09-02 (plan A2 pre-check):
# the source's season line divided by 17 x Sleeper's week-1 line for five
# healthy starters (Gibbs, Allen, Nacua, Bowers, B. Robinson). The sheet and
# ESPN sit at ~0.98 -- full-season totals; Sleeper/Rotowire at ~0.92 -- its
# season line already embeds about one missed game. A source that already
# discounts games is EXCLUDED from the games table's per-row scale (it keeps
# the uniform `games`), so nobody is discounted twice. Five players is thin;
# re-measure when the table is re-derived.
SOURCE_GAMES_CONVENTION = {
    "fantasypros_sheet": {"ratio": 0.98, "already_discounted": False},
    "espn_projections": {"ratio": 0.98, "already_discounted": False},
    "sleeper_rotowire": {"ratio": 0.92, "already_discounted": True},
}
DISCOUNTED_SOURCES = tuple(k for k, v in SOURCE_GAMES_CONVENTION.items() if v["already_discounted"])
SCHEMA = {"sleeper_id": pl.Utf8, "name": pl.Utf8, "pos": pl.Utf8, "team": pl.Utf8,
          "pts17": pl.Float64, "source": pl.Utf8, "as_of": pl.Utf8, "line": pl.Utf8}
# what combine() emits: the schema plus the dispersion across sources (plan A1)
# `pts17_sd` is DISAGREEMENT BETWEEN SOURCES and `n_sources` counts them.
# `pts17_band` is a different quantity: ONE source's own stated uncertainty,
# from a high/low line it publishes alongside its base line. Conflating the
# two would make pts17_sd mean two things and would quietly defeat the
# n_sources >= 2 guard, so the band gets its own column and is null for a
# source that publishes no range.
DISPERSION = {"n_sources": pl.Int64, "pts17_sd": pl.Float64, "pts17_hi": pl.Float64,
              "pts17_lo": pl.Float64, "pts17_band": pl.Float64}
SCHEMA_COMBINED = {**SCHEMA, **DISPERSION}

# Column layout of each sheet position tab (0-based, after Player, Team).
# The header names repeat ("YDS" twice), so the mapping is positional.
SHEET_COLS = {
    "QB": ["pass_att", "pass_cmp", "pass_yd", "pass_td", "pass_int",
           "rush_att", "rush_yd", "rush_td", "fum_lost"],
    "RB": ["rush_att", "rush_yd", "rush_td", "rec", "rec_yd", "rec_td", "fum_lost"],
    "WR": ["rec", "rec_yd", "rec_td", "rush_att", "rush_yd", "rush_td", "fum_lost"],
    "TE": ["rec", "rec_yd", "rec_td", "fum_lost"],
}


def empty() -> pl.DataFrame:
    return pl.DataFrame(schema=SCHEMA)


def _frame(rows: list[dict], schema: dict | None = None) -> pl.DataFrame:
    schema = schema or SCHEMA
    return pl.DataFrame(rows, schema=schema) if rows else pl.DataFrame(schema=schema)


# ------------------------------------------------------------------ sheet

def sheet_bump_column(ws_formulas) -> int | None:
    """0-based index of the tab's ROOKIE BUMP column, located by formula shape.

    The sheet adds a per-position rookie adjustment to every projection:

        ppg   = raw scored line / 17
        bump  = IF(rookie, MAX(0, k * (cap - ppg)), 0)     per game
        total = raw scored line + bump * 17

    with k/cap of 0.258/14.9 at RB and 0.28/12.0 at WR, and no bump at QB or
    TE. Reading Excel's own cached product (the `= <T> * 17` column) rather
    than re-deriving it from those constants means a re-published sheet with
    different coefficients is followed automatically, and there is no second
    copy of somebody else's numbers to go stale in this repo.
    """
    for col in range(1, (ws_formulas.max_column or 0) + 1):
        v = ws_formulas.cell(row=3, column=col).value
        if isinstance(v, str) and re.fullmatch(r"=[A-Z]+\d+\*17", v.strip()):
            return col - 1
    return None


def parse_sheet_tab(rows: list[tuple], pos: str, bump_col: int | None = None) -> list[dict]:
    """A player row carries the name; the 'high' and 'low' rows that follow
    are the same expert's own range for him and are ATTACHED to him.

    They used to be skipped. The sheet's own Aggregate tab averages low, base
    and high, so throwing two of the three away meant the loader never
    reproduced the number the spreadsheet itself reports.

    Returns [{name, team, line, line_hi, line_lo, bump}]: the extremes are
    None when absent, and `bump` is the tab's rookie adjustment in season
    points (0.0 for a veteran, None when the column was not located).
    """
    cols = SHEET_COLS[pos]
    out: list[dict] = []
    for r in rows[1:]:
        name = r[0] if len(r) > 0 else None
        named = isinstance(name, str) and name.replace("\xa0", "").replace("Â", "").strip()
        marker = str(r[1]).strip().lower() if len(r) > 1 and isinstance(r[1], str) else ""
        line = {k: float(r[2 + i]) for i, k in enumerate(cols)
                if len(r) > 2 + i and isinstance(r[2 + i], (int, float))}
        if not named:
            # an unnamed row marked high/low belongs to the player above it;
            # anything else is the spacer row (a non-breaking space)
            if marker in ("high", "low") and out and line:
                out[-1][f"line_{marker[:2]}"] = line
            continue
        bump = None
        if bump_col is not None and len(r) > bump_col:
            b = r[bump_col]
            bump = float(b) if isinstance(b, (int, float)) else 0.0
        out.append({"name": name.strip(), "team": r[1] if len(r) > 1 else None,
                    "line": line, "line_hi": None, "line_lo": None, "bump": bump})
    return out


def from_sheet(path: Path, scoring: dict, index, as_of: str) -> tuple[pl.DataFrame, list[str]]:
    """The sheet's consensus lines in the common schema. `index` is a
    SleeperIndex (name, pos, team -> sleeper_id). Returns (frame, unmatched)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # a second pass for FORMULAS, only to locate the rookie-bump column: the
    # values pass cannot see formulas and the column sits at no fixed letter
    wf = openpyxl.load_workbook(path, read_only=True, data_only=False)
    rows, unmatched, bumped = [], [], 0
    for pos in SHEET_COLS:
        bump_col = sheet_bump_column(wf[pos])
        for p in parse_sheet_tab(list(wb[pos].iter_rows(values_only=True)), pos, bump_col):
            sid = index.match(p["name"], pos, p["team"])
            if not sid:
                unmatched.append(f"{p['name']} ({pos})")
                continue
            # the sheet's own number is the scored line PLUS its rookie bump;
            # reading only the line under-projects every rookie by up to 65
            # points against the spreadsheet this source exists to carry
            bump = float(p.get("bump") or 0.0)
            if bump:
                bumped += 1
            base = float(score_projection(p["line"], scoring)) + bump
            # the source's own range, as a one-sigma-equivalent. Population sd
            # of {low, base, high} -- the SAME estimator combine(mode="mean")
            # uses across sources, so the two dispersion numbers are at least
            # on one scale even though they measure different things.
            # the sheet adds the same bump to low, base and high, so it
            # shifts the trio without widening it: the band is unchanged
            trio = [base]
            for k in ("line_lo", "line_hi"):
                if p.get(k):
                    trio.append(float(score_projection(p[k], scoring)) + bump)
            band = None
            if len(trio) == 3:
                mu = sum(trio) / 3.0
                band = (sum((x - mu) ** 2 for x in trio) / 3.0) ** 0.5
            rows.append({"sleeper_id": str(sid), "name": p["name"], "pos": pos, "team": p["team"],
                         "pts17": base,
                         "source": "fantasypros_sheet", "as_of": as_of,
                         "line": json.dumps(p["line"], sort_keys=True),
                         "pts17_band": band})
    if bumped:
        log.info("sheet: rookie bump applied to %d players", bumped)
    return (_frame(rows, {**SCHEMA, "pts17_band": pl.Float64})
            .unique(subset="sleeper_id", keep="first"), unmatched)


# ---------------------------------------------------------------- sleeper

def from_sleeper(season: int, scoring: dict, raw_dir: Path, getter=None,
                 ttl: int | None = None) -> pl.DataFrame:
    """Sleeper's season stat lines in the common schema (rows with no stat
    line beyond ADP placeholders are unprojected and left out)."""
    rows = []
    kw = {}
    if getter is not None:
        kw["getter"] = getter
    if ttl is not None:
        kw["ttl"] = ttl
    for pos in SLEEPER_POSITIONS:
        for r in fetch_position(season, pos, raw_dir, **kw):
            stats = r.get("stats") or {}
            line = {k: v for k, v in stats.items()
                    if not k.startswith("adp_") and not k.startswith("pos_adp") and k != "gp"
                    and v is not None}
            if not line:
                continue
            p = r.get("player") or {}
            upd = r.get("updated_at") or r.get("last_modified")
            as_of = (dt.datetime.fromtimestamp(upd / 1000, tz=dt.timezone.utc).date().isoformat()
                     if upd else "")
            rows.append({"sleeper_id": str(r.get("player_id")),
                         "name": " ".join(x for x in (p.get("first_name"), p.get("last_name")) if x) or None,
                         "pos": p.get("position") or pos, "team": r.get("team"),
                         "pts17": float(score_projection(line, scoring)),
                         "source": "sleeper_rotowire", "as_of": as_of,
                         "line": json.dumps(line, sort_keys=True)})
    return _frame(rows).unique(subset="sleeper_id", keep="first")


# ------------------------------------------------------------------- espn

def from_espn(season: int, scoring: dict, raw_dir: Path, id_map: pl.DataFrame, index,
              getter=None, ttl: int | None = None) -> tuple[pl.DataFrame, list[str]]:
    """ESPN's season lines in the common schema (plan A1). Ids resolve through
    the id map's espn_id first, then the same name matcher the market table
    uses; unmatched names are RETURNED. Team is left to the board (the feed
    carries a numeric proTeamId)."""
    from . import espn as E
    from .market import _attach_sleeper_ids
    kw = {}
    if getter is not None:
        kw["getter"] = getter
    if ttl is not None:
        kw["ttl"] = ttl
    parsed = E.parse_players(E.fetch_projections(season, raw_dir, **kw), season)
    if not parsed:
        return empty(), []
    df = pl.DataFrame({"espn_id": [p["espn_id"] for p in parsed], "name": [p["name"] for p in parsed],
                       "pos": [p["pos"] for p in parsed], "team": [None] * len(parsed)},
                      schema={"espn_id": pl.Utf8, "name": pl.Utf8, "pos": pl.Utf8, "team": pl.Utf8})
    df, unmatched = _attach_sleeper_ids(df, index, id_map, via_fp_id=True, id_col="espn_id")
    as_of = E.cache_as_of(raw_dir, season)
    rows = []
    for p, sid in zip(parsed, df["sleeper_id"].to_list()):
        if not sid:
            continue
        rows.append({"sleeper_id": str(sid), "name": p["name"], "pos": p["pos"], "team": None,
                     "pts17": float(score_projection(p["line"], scoring)),
                     "source": "espn_projections", "as_of": as_of,
                     "line": json.dumps(p["line"], sort_keys=True)})
    return _frame(rows).unique(subset="sleeper_id", keep="first"), unmatched


# ------------------------------------------------------------------ union

def _with_dispersion_single(f: pl.DataFrame) -> pl.DataFrame:
    """One source: no cross-source disagreement by construction. Its own band
    is carried through untouched when it published one."""
    if "pts17_band" not in f.columns:
        f = f.with_columns(pl.lit(None, dtype=pl.Float64).alias("pts17_band"))
    f = f.with_columns(pl.lit(1, dtype=pl.Int64).alias("n_sources"), pl.lit(0.0).alias("pts17_sd"),
                       pl.col("pts17").alias("pts17_hi"), pl.col("pts17").alias("pts17_lo"))
    # a source that already carries pts17_band leaves it mid-frame, so the
    # canonical order is restored explicitly rather than depending on which
    # columns each source happened to supply
    return f.select(list(SCHEMA_COMBINED))


def combine(frames: list[pl.DataFrame], mode: str = "first", scoring: dict | None = None) -> pl.DataFrame:
    """mode 'first': first source wins per player; later sources fill the
    gaps (the 2026 default; byte-identical to before plan A1, plus the four
    dispersion columns at their single-source values).
    mode 'mean': the equal-weight per-stat MEAN of every source that carries
    the player, scored ONCE (`scoring` required); a stat a source omits
    counts as 0 for that source. Because scoring is linear the scored mean
    equals the mean of the per-source pts17, so the dispersion columns are
    the population std / max / min of the per-source scores."""
    frames = [f for f in frames if f.height]
    if mode == "first":
        out = pl.DataFrame(schema=SCHEMA_COMBINED)
        for f in frames:
            add = f.filter(~pl.col("sleeper_id").is_in(out["sleeper_id"].to_list())) if out.height else f
            add = _with_dispersion_single(add)
            out = pl.concat([out, add], how="vertical") if out.height else add
        return out
    if mode != "mean":
        raise ValueError(f"unknown combine mode {mode!r}")
    if scoring is None:
        raise ValueError("combine(mode='mean') needs the league scoring")
    if not frames:
        return pl.DataFrame(schema=SCHEMA_COMBINED)
    allrows = pl.concat(frames, how="vertical")
    out = []
    for sid, g in allrows.group_by("sleeper_id", maintain_order=True):
        sid = sid[0] if isinstance(sid, tuple) else sid
        lines = [json.loads(x) for x in g["line"].to_list()]
        n = len(lines)
        keys = sorted({k for ln in lines for k in ln})
        mean_line = {k: sum(float(ln.get(k, 0.0)) for ln in lines) / n for k in keys}
        scores = [float(score_projection(ln, scoring)) for ln in lines]
        mu = sum(scores) / n
        sd = (sum((s - mu) ** 2 for s in scores) / n) ** 0.5
        first = g.row(0, named=True)
        out.append({"sleeper_id": str(sid),
                    "name": next((x for x in g["name"].to_list() if x), first["name"]),
                    "pos": next((x for x in g["pos"].to_list() if x), first["pos"]),
                    "team": next((x for x in g["team"].to_list() if x), None),
                    "pts17": float(score_projection(mean_line, scoring)),
                    "source": "mean(" + ",".join(sorted(set(g["source"].to_list()))) + ")",
                    "as_of": max((x for x in g["as_of"].to_list() if x), default=""),
                    "line": json.dumps(mean_line, sort_keys=True),
                    "n_sources": n, "pts17_sd": sd, "pts17_hi": max(scores), "pts17_lo": min(scores),
                    # the mean of the sources that published a range. Combining
                    # a within-source band with cross-source disagreement (in
                    # quadrature, say) is a modelling choice and is NOT made
                    # here: the two stay separate columns.
                    "pts17_band": (lambda b: sum(b) / len(b) if b else None)(
                        [float(x) for x in g["pts17_band"].to_list() if x is not None]
                        if "pts17_band" in g.columns else [])})
    return pl.DataFrame(out, schema=SCHEMA_COMBINED)


def load_external(cfg, index, getter=None) -> tuple[pl.DataFrame, dict]:
    """The configured sources, in order, in the common schema. Report says
    what each contributed and what the sheet could not match."""
    p = cfg.get("projections") or {}
    ext = p.get("external") or {}
    scoring = {k: float(v) for k, v in (cfg.get("scoring") or (cfg.get("expected") or {}).get("scoring") or {}).items()}
    if not scoring:
        raise ValueError("league yaml carries no scoring block")
    mode = str(ext.get("combine", "first"))
    frames, report = [], {"sources": [], "sheet_unmatched": [], "espn_unmatched": [], "combine": mode}
    for name in ext.get("sources") or ["sleeper"]:
        if name == "sheet":
            path = Path(cfg.root) / str(ext.get("sheet_path", ""))
            if not path.exists():
                report["sources"].append({"source": "sheet", "rows": 0, "error": f"missing {path.name}"})
                continue
            f, unmatched = from_sheet(path, scoring, index, as_of=str(ext.get("sheet_as_of", "")))
            report["sheet_unmatched"] = unmatched
        elif name == "sleeper":
            try:
                f = from_sleeper(int(cfg["season"]), scoring, cfg.path("raw"), getter=getter)
            except ConsensusUnavailable as e:
                report["sources"].append({"source": "sleeper", "rows": 0, "error": str(e)})
                continue
        elif name == "espn":
            from .espn import EspnUnavailable
            from .ids import load_id_map
            try:
                f, unmatched = from_espn(int(cfg["season"]), scoring, cfg.path("raw"),
                                         load_id_map(cfg.path("raw")), index)
                report["espn_unmatched"] = unmatched
            except EspnUnavailable as e:
                report["sources"].append({"source": "espn", "rows": 0, "error": str(e)})
                continue
        else:
            raise ValueError(f"unknown projection source {name!r}")
        report["sources"].append({"source": name, "rows": f.height,
                                  "as_of": (f["as_of"].drop_nulls().max() if f.height else None)})
        frames.append(f)
    out = combine(frames, mode=mode, scoring=scoring)
    report["total"] = out.height
    return out, report


# -------------------------------------------------------- non-starter rule

def zero_non_starters(df: pl.DataFrame, depth: pl.DataFrame, teams: int,
                      starters: dict | None = None) -> pl.DataFrame:
    """proj_pts -> 0 for players the depth chart lists behind the starters
    AND the market does not rank as a starter (or does not rank at all).
    Adds `contingent_of` (the order-1 player at the same team/position) and
    `non_starter` (bool). df needs sleeper_id, pos, team, ecr, adp, proj_pts.

    Positions with a single ordered depth chart only (role.GATED: QB, RB,
    TE). Sleeper's receiver chart is per slot (LWR/RWR/SWR), so a WR "order"
    is not an overall depth and WR is left alone."""
    st = starters or STARTERS
    dcols = ["sleeper_id", "depth_order"] + (["depth_pos"] if "depth_pos" in depth.columns else [])
    d = df.join(depth.select(dcols), on="sleeper_id", how="left")
    if "depth_pos" not in d.columns:
        d = d.with_columns(pl.col("pos").alias("depth_pos"))
    d = d.with_columns(
        pl.coalesce(pl.col("ecr"), pl.col("adp")).rank(method="ordinal").over("pos").alias("_mkt_rank"))
    starters_expr = pl.col("pos").replace_strict(st, default=None, return_dtype=pl.Int64)
    depth_backup = (pl.col("pos").is_in(list(GATED)) & (pl.col("depth_pos") == pl.col("pos"))
                    & pl.col("depth_order").is_not_null() & (pl.col("depth_order") > starters_expr))
    market_backup = pl.col("_mkt_rank").is_null() | (pl.col("_mkt_rank") > starters_expr * teams)
    gate = depth_backup & market_backup
    # the starter he is contingent on: order 1 at the same team and chart position
    ones = (d.filter((pl.col("depth_order") == 1) & (pl.col("depth_pos") == pl.col("pos")))
              .select("team", "pos", pl.col("name").alias("contingent_of"))
              .unique(subset=["team", "pos"], keep="first"))
    d = d.join(ones, on=["team", "pos"], how="left")
    d = d.with_columns(
        gate.alias("non_starter"),
        pl.when(gate).then(pl.col("contingent_of")).otherwise(None).alias("contingent_of"),
        pl.when(gate & pl.col("proj_pts").is_not_null()).then(0.0).otherwise(pl.col("proj_pts")).alias("proj_pts"),
    )
    return d.drop("_mkt_rank", "depth_order", "depth_pos")


def depth_table(raw_dir: Path) -> pl.DataFrame | None:
    return depth_orders(raw_dir)
